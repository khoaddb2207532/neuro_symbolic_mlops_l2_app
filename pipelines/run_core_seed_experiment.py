"""Chạy hoặc tiếp tục thí nghiệm lõi MobileNetV3-Small + GFlowNet-DB.

Module này thay thế logic orchestration trong notebook Kaggle:

* cấu hình một seed;
* tùy chọn khôi phục artefact từ Kaggle Add Input;
* tự bỏ qua các stage đã hoàn tất;
* dùng matched budget bằng đúng số luật GFlowNet chọn;
* kiểm tra fairness và xuất CSV chuẩn hóa cho notebook tổng hợp.

Ví dụ:
    python -m pipelines.run_core_seed_experiment \
        --config params.yaml \
        --seed 42 \
        --backbone mobilenetv3_small \
        --data-dir /kaggle/input/.../vietnamese_cultural_dataset \
        --output-dir /kaggle/working/neuro_symbolic_mlops_l2_app/outputs_seed_42 \
        --kaggle-input-root /kaggle/input
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import tarfile
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import yaml
from openpyxl import load_workbook


HEURISTICS = ("random", "topk_confidence", "greedy_coverage")
SUPPORTED_BACKBONES = (
    "mobilenetv3_small",
    "resnet50",
    "densenet121",
    "efficientnet_b0",
    "swin_t",
    "vit_b_16",
)


def _safe_identifier(value: str) -> str:
    normalized = "".join(
        character.lower()
        if character.isalnum() or character in {"-", "_"}
        else "-"
        for character in value.strip()
    )
    normalized = "-".join(filter(None, normalized.split("-")))
    if not normalized:
        raise ValueError("Identifier không được rỗng.")
    return normalized


def _dataset_fingerprint(data_dir: Path) -> str:
    """Hash metadata cây dữ liệu, không đọc toàn bộ nội dung ảnh."""
    digest = hashlib.sha256()
    files = sorted(path for path in data_dir.rglob("*") if path.is_file())
    for path in files:
        stat = path.stat()
        relative = path.relative_to(data_dir).as_posix()
        digest.update(relative.encode("utf-8"))
        digest.update(b"\0")
        digest.update(str(stat.st_size).encode("ascii"))
        digest.update(b"\n")
    digest.update(f"file_count={len(files)}".encode("ascii"))
    return digest.hexdigest()


def _git_commit(project: Path) -> str:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "HEAD"],
            cwd=project,
            text=True,
        ).strip()
    except (subprocess.CalledProcessError, FileNotFoundError):
        return "unknown"


def _write_run_manifest(path: Path, manifest: Dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

_RUNTIME_EVENTS: List[Dict] = []
_RUNTIME_LEDGER_PATH: Optional[Path] = None
_RUNTIME_CONTEXT: Dict = {}


def _runtime_stage_name(module: str, args: Tuple[object, ...]) -> str:
    mapping = {
        "pipelines.train_all_baselines": "baseline_cnn",
        "pipelines.stage2_extract_features": "feature_extraction",
        "pipelines.stage3_extract_rules": "rule_extraction",
        "pipelines.stage4_select_rules_gflownet": "gflownet_stage4",
        "pipelines.stage4b_analyze_rule_rankings": "ranking_analysis_stage4b",
        "pipelines.stage4c_evaluate_gflownet_checkpoints": (
            "checkpoint_evaluation_stage4c"
        ),
        "pipelines.stage5_train_rule_regularized": (
            "gflownet_fixed_stage5"
        ),
        "pipelines.stage5_train_rule_bayesian": "bayesian_stage5",
        "pipelines.evaluate_saved_checkpoints": "exact_checkpoint_evaluation",
        "pipelines.analyze_rule_set_quality": "rule_set_quality_analysis",
    }
    if module == "pipelines.stage5_train_rule_regularized_heuristics":
        string_args = list(map(str, args))
        method = (
            string_args[string_args.index("--methods") + 1]
            if "--methods" in string_args
            else "unknown"
        )
        return f"heuristic_{method}_selection_and_stage5"
    return mapping.get(module, module)


def _persist_runtime_events() -> None:
    if _RUNTIME_LEDGER_PATH is None:
        return
    _RUNTIME_LEDGER_PATH.parent.mkdir(parents=True, exist_ok=True)
    _RUNTIME_LEDGER_PATH.write_text(
        json.dumps(_RUNTIME_EVENTS, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _initialize_runtime_tracking(
    output_dir: Path,
    *,
    seed: int,
    backbone: str,
) -> None:
    global _RUNTIME_EVENTS, _RUNTIME_LEDGER_PATH, _RUNTIME_CONTEXT
    _RUNTIME_LEDGER_PATH = output_dir / "runtime_events.json"
    if _RUNTIME_LEDGER_PATH.exists():
        _RUNTIME_EVENTS = json.loads(
            _RUNTIME_LEDGER_PATH.read_text(encoding="utf-8")
        )
    else:
        _RUNTIME_EVENTS = []
    _RUNTIME_CONTEXT = {"seed": seed, "backbone": backbone}


def _run_module(project: Path, config_path: Path, module: str, *args: object) -> None:
    command = [
        sys.executable,
        "-m",
        module,
        "--config",
        str(config_path),
        *map(str, args),
    ]
    print("\n$", " ".join(command), flush=True)
    stage = _runtime_stage_name(module, args)
    started_at = datetime.now(timezone.utc)
    timer = time.perf_counter()
    status = "completed"
    return_code = 0
    try:
        subprocess.run(command, cwd=project, check=True)
    except subprocess.CalledProcessError as error:
        status = "failed"
        return_code = int(error.returncode)
        raise
    finally:
        elapsed_seconds = time.perf_counter() - timer
        _RUNTIME_EVENTS.append(
            {
                **_RUNTIME_CONTEXT,
                "stage": stage,
                "module": module,
                "status": status,
                "return_code": return_code,
                "started_at_utc": started_at.isoformat(),
                "finished_at_utc": datetime.now(timezone.utc).isoformat(),
                "elapsed_seconds": elapsed_seconds,
                "elapsed_minutes": elapsed_seconds / 60.0,
                "command_args": list(map(str, args)),
            }
        )
        _persist_runtime_events()


def _write_config(
    config_path: Path,
    *,
    seed: int,
    backbone: str,
    data_dir: Path,
    output_dir: Path,
    selection_budget: Optional[int] = None,
) -> Dict:
    config = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    config["seed"] = seed
    config["data_dir"] = str(data_dir)
    config["output_dir"] = str(output_dir)
    config["batch_size"] = 64
    config["num_workers"] = 2
    config["num_epochs"] = 100
    config["patience"] = 5
    config["baseline_comparison"]["selected_architecture"] = backbone
    config["baseline_comparison"]["architectures"] = [backbone]
    config["gflownet"]["loss_type"] = "db"
    if selection_budget is None:
        config.pop("selection_budget", None)
    else:
        config["selection_budget"] = int(selection_budget)
    config_path.write_text(
        yaml.safe_dump(config, sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )
    return config


def _safe_extract_tar(archive: Path, destination: Path) -> None:
    destination = destination.resolve()
    destination.mkdir(parents=True, exist_ok=True)
    with tarfile.open(archive, "r:gz") as tar:
        for member in tar.getmembers():
            target = (destination / member.name).resolve()
            if target != destination and destination not in target.parents:
                raise RuntimeError(f"Archive chứa path không an toàn: {member.name}")
        tar.extractall(destination)


def _find_restorable_output(input_root: Path, seed: int) -> Tuple[str, Path] | None:
    directories = sorted(
        path
        for path in input_root.rglob(f"outputs_seed_{seed}")
        if path.is_dir()
    )
    if directories:
        return "directory", directories[0]

    archives = sorted(input_root.rglob(f"seed_{seed}_artifacts.tar.gz"))
    if archives:
        return "archive", archives[0]
    return None


def _validate_output_identity(
    output_dir: Path,
    seed: int,
    backbone: str,
    dataset_id: Optional[str] = None,
) -> None:
    metadata_path = output_dir / "experiment_metadata.json"
    if metadata_path.exists():
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        if int(metadata.get("seed", -1)) != seed or metadata.get("backbone") != backbone:
            raise ValueError(
                "Kaggle Input không khớp thí nghiệm yêu cầu: "
                f"metadata={metadata}, requested={{'seed': {seed}, "
                f"'backbone': '{backbone}'}}."
            )
        stored_dataset = metadata.get("dataset_id")
        if dataset_id and stored_dataset and stored_dataset != dataset_id:
            raise ValueError(
                "Kaggle Input thuộc dataset khác: "
                f"stored={stored_dataset}, requested={dataset_id}."
            )
        return

    # Tương thích output cũ chưa có metadata: suy ra backbone từ thư mục
    # baseline. Không cho phép dùng feature/rule của backbone khác.
    baseline_root = output_dir / "baseline_comparison"
    model_dirs = sorted(
        path.name for path in baseline_root.iterdir() if path.is_dir()
    ) if baseline_root.is_dir() else []
    if model_dirs and backbone not in model_dirs:
        raise ValueError(
            f"Output restore chứa backbone {model_dirs}, không phải {backbone}. "
            "Hãy Add Input đúng notebook/backbone."
        )


def _restore_if_available(
    input_root: Optional[Path],
    output_dir: Path,
    seed: int,
    backbone: str,
    dataset_id: Optional[str] = None,
) -> None:
    if input_root is None or not input_root.is_dir():
        return
    source = _find_restorable_output(input_root, seed)
    if source is None:
        print(
            f"Không tìm thấy output seed {seed} trong {input_root}; "
            "bắt đầu/chạy tiếp bằng artefact hiện có.",
            flush=True,
        )
        return

    kind, path = source
    print(f"Khôi phục output seed {seed} từ {path}", flush=True)
    if kind == "directory":
        shutil.copytree(path, output_dir, dirs_exist_ok=True)
    else:
        _safe_extract_tar(path, output_dir)
    _validate_output_identity(output_dir, seed, backbone, dataset_id)


def _restore_bayesian_if_available(
    input_root: Optional[Path],
    output_dir: Path,
    seed: int,
    backbone: str,
    dataset_id: Optional[str] = None,
) -> Optional[Path]:
    """Khôi phục Bayesian Stage 5 từ output notebook cũ nếu có.

    Output Bayesian cũ có thể nằm trong một ``outputs_seed_<seed>`` khác với
    output lõi, hoặc chỉ còn archive ``seed_<seed>_bayesian_artifacts.tar.gz``.
    Chỉ classification report được xem là completion marker.
    """
    destination = output_dir / "05b_rules_model_bayesian"
    existing_report = _first_report(destination)
    if existing_report is not None:
        return existing_report
    if input_root is None or not input_root.is_dir():
        return None

    # Ưu tiên cây output đầy đủ vì nó giữ cả metadata seed/backbone.
    for candidate_root in sorted(
        path
        for path in input_root.rglob(f"outputs_seed_{seed}")
        if path.is_dir()
    ):
        candidate_bayesian = (
            candidate_root / "05b_rules_model_bayesian"
        )
        candidate_report = _first_report(candidate_bayesian)
        if candidate_report is None:
            continue
        _validate_output_identity(
            candidate_root, seed, backbone, dataset_id
        )
        print(
            "Khôi phục Bayesian Stage 5 từ output notebook cũ:",
            candidate_bayesian,
        )
        shutil.copytree(
            candidate_bayesian,
            destination,
            dirs_exist_ok=True,
        )
        restored_report = _first_report(destination)
        if restored_report is not None:
            return restored_report

    # Fallback cho notebook cũ chỉ publish archive Bayesian riêng.
    for archive in sorted(
        input_root.rglob(f"seed_{seed}_bayesian_artifacts.tar.gz")
    ):
        print("Khôi phục Bayesian Stage 5 từ archive cũ:", archive)
        _safe_extract_tar(archive, destination)
        restored_report = _first_report(destination)
        if restored_report is not None:
            return restored_report
        print(
            "Archive không chứa classification report hoàn tất, "
            "tiếp tục tìm candidate khác."
        )
    return None


def _xlsx_row_count(path: Path) -> int:
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        # File save_rules_excel có một hàng header.
        return max(sheet.max_row - 1, 0)
    finally:
        workbook.close()


def _first_report(directory: Path) -> Optional[Path]:
    reports = sorted(directory.glob("*classification_report.txt"))
    return reports[0] if reports else None


def _report_metrics(path: Path) -> Tuple[float, float]:
    text = path.read_text(encoding="utf-8")
    accuracy_match = re.search(r"Overall Accuracy:\s*([0-9.]+)%", text)
    if accuracy_match is None:
        raise ValueError(f"Không đọc được accuracy từ {path}")
    macro_line = next(
        (line for line in text.splitlines() if line.strip().startswith("macro avg")),
        None,
    )
    if macro_line is None:
        raise ValueError(f"Không đọc được macro-F1 từ {path}")
    return float(accuracy_match.group(1)) / 100.0, float(macro_line.split()[4])


def _required_dataset_splits(data_dir: Path) -> None:
    missing = [name for name in ("train", "val", "test") if not (data_dir / name).is_dir()]
    if missing:
        raise FileNotFoundError(f"{data_dir} thiếu các split: {missing}")


def _write_csv(path: Path, rows: Iterable[Dict], fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_runtime_summary(
    output_dir: Path,
    working_dir: Path,
    *,
    seed: int,
    backbone: str,
    include_bayesian: bool,
) -> Tuple[Path, Path]:
    expected_stages = [
        "baseline_cnn",
        "feature_extraction",
        "rule_extraction",
        "gflownet_stage4",
        "gflownet_fixed_stage5",
        "heuristic_random_selection_and_stage5",
        "heuristic_topk_confidence_selection_and_stage5",
        "heuristic_greedy_coverage_selection_and_stage5",
    ]
    if include_bayesian:
        expected_stages.append("bayesian_stage5")

    rows = []
    for stage in expected_stages:
        completed = [
            event
            for event in _RUNTIME_EVENTS
            if event.get("stage") == stage
            and event.get("status") == "completed"
        ]
        if completed:
            event = completed[-1]
            rows.append(
                {
                    "seed": seed,
                    "backbone": backbone,
                    "stage": stage,
                    "runtime_status": "measured",
                    "elapsed_seconds": event["elapsed_seconds"],
                    "elapsed_minutes": event["elapsed_minutes"],
                    "started_at_utc": event["started_at_utc"],
                    "finished_at_utc": event["finished_at_utc"],
                }
            )
        else:
            rows.append(
                {
                    "seed": seed,
                    "backbone": backbone,
                    "stage": stage,
                    "runtime_status": "unavailable_historical",
                    "elapsed_seconds": "",
                    "elapsed_minutes": "",
                    "started_at_utc": "",
                    "finished_at_utc": "",
                }
            )

    summary_source = output_dir / "runtime_summary.csv"
    _write_csv(summary_source, rows, list(rows[0]))
    summary_destination = (
        working_dir / f"seed_{seed}_runtime_summary.csv"
    )
    events_destination = (
        working_dir / f"seed_{seed}_runtime_events.json"
    )
    shutil.copy2(summary_source, summary_destination)
    if _RUNTIME_LEDGER_PATH is None or not _RUNTIME_LEDGER_PATH.exists():
        raise FileNotFoundError("Runtime event ledger chưa được tạo.")
    shutil.copy2(_RUNTIME_LEDGER_PATH, events_destination)
    return summary_destination, events_destination


def run(args: argparse.Namespace) -> None:
    project = Path(args.project_dir).resolve()
    config_path = Path(args.config)
    if not config_path.is_absolute():
        config_path = project / config_path
    data_dir = Path(args.data_dir)
    output_dir = Path(args.output_dir)
    working_dir = Path(args.working_dir)
    input_root = Path(args.kaggle_input_root) if args.kaggle_input_root else None
    dataset_id = _safe_identifier(args.dataset_id)
    account_id = _safe_identifier(args.account_id)
    run_id = _safe_identifier(
        args.run_id
        or f"{dataset_id}__{args.backbone}__db__seed_{args.seed}"
    )

    _required_dataset_splits(data_dir)
    dataset_fingerprint = _dataset_fingerprint(data_dir)
    git_commit = _git_commit(project)
    output_dir.mkdir(parents=True, exist_ok=True)
    _restore_if_available(
        input_root,
        output_dir,
        args.seed,
        args.backbone,
        dataset_id,
    )
    _initialize_runtime_tracking(
        output_dir,
        seed=args.seed,
        backbone=args.backbone,
    )
    manifest_path = output_dir / "run_manifest.json"
    run_manifest = {
        "schema_version": 1,
        "run_id": run_id,
        "dataset_id": dataset_id,
        "dataset_path": str(data_dir),
        "dataset_fingerprint": dataset_fingerprint,
        "seed": args.seed,
        "backbone": args.backbone,
        "gflownet_loss": "db",
        "account_id": account_id,
        "git_commit": git_commit,
        "status": "running",
        "started_at_utc": datetime.now(timezone.utc).isoformat(),
        "expected_methods": [
            "cnn_baseline",
            "gflownet_db",
            "gflownet_db_bayesian",
            "random",
            "topk_confidence",
            "greedy_coverage",
        ],
    }
    _write_run_manifest(manifest_path, run_manifest)
    config = _write_config(
        config_path,
        seed=args.seed,
        backbone=args.backbone,
        data_dir=data_dir,
        output_dir=output_dir,
    )
    (output_dir / "experiment_metadata.json").write_text(
        json.dumps(
            {
                "run_id": run_id,
                "dataset_id": dataset_id,
                "dataset_fingerprint": dataset_fingerprint,
                "seed": args.seed,
                "backbone": args.backbone,
                "loss_type": "db",
                "git_commit": git_commit,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    baseline_dir = output_dir / "baseline_comparison" / args.backbone
    baseline_checkpoint = baseline_dir / "baseline_best.pth"
    baseline_report = _first_report(baseline_dir)
    if baseline_checkpoint.exists() and baseline_report is not None:
        print("SKIP baseline:", baseline_checkpoint)
    else:
        print(
            "RESUME baseline:",
            f"checkpoint={'OK' if baseline_checkpoint.exists() else 'MISSING'},",
            f"report={'OK' if baseline_report is not None else 'MISSING'}",
        )
        _run_module(
            project,
            config_path,
            "pipelines.train_all_baselines",
            "--models",
            args.backbone,
            "--epochs",
            config["num_epochs"],
        )
        baseline_report = _first_report(baseline_dir)
        if not baseline_checkpoint.exists() or baseline_report is None:
            raise FileNotFoundError(
                "Baseline chạy xong nhưng thiếu checkpoint/report."
            )

    features_dir = output_dir / "02_features"
    required_features = [
        features_dir / f"{split}_{kind}.pt"
        for split in ("train", "val", "test")
        for kind in ("features", "labels")
    ]
    missing_features = [path for path in required_features if not path.exists()]
    if not missing_features:
        print("SKIP Stage 2: feature artefact đã tồn tại.")
    else:
        print(
            "RESUME Stage 2, thiếu:",
            ", ".join(path.name for path in missing_features),
        )
        _run_module(project, config_path, "pipelines.stage2_extract_features")
        missing_features = [path for path in required_features if not path.exists()]
        if missing_features:
            raise FileNotFoundError(
                "Stage 2 chạy xong nhưng còn thiếu: "
                + ", ".join(map(str, missing_features))
            )

    raw_rules = output_dir / "03_rules" / "raw_rules.pkl"
    if raw_rules.exists():
        print("SKIP Stage 3:", raw_rules)
    else:
        _run_module(project, config_path, "pipelines.stage3_extract_rules")

    gfn_rules_dir = output_dir / "04_filtered_rules"
    gfn_rules_xlsx = gfn_rules_dir / "selected_rules.xlsx"
    gfn_rules_pickle = gfn_rules_dir / "selected_rules.pkl"
    gfn_rule_order = gfn_rules_dir / "gflownet_rule_order.pkl"
    gfn_diverse_checkpoint = (
        gfn_rules_dir / "gflownet_best_diverse.pth"
    )
    gfn_converged_checkpoint = (
        gfn_rules_dir / "gflownet_best_converged.pth"
    )
    has_gfn_checkpoint = (
        gfn_diverse_checkpoint.exists()
        or gfn_converged_checkpoint.exists()
    )
    if (
        gfn_rules_xlsx.exists()
        and gfn_rules_pickle.exists()
        and gfn_rule_order.exists()
        and has_gfn_checkpoint
    ):
        print("SKIP Stage 4:", gfn_rules_xlsx)
    else:
        print(
            "RESUME Stage 4:",
            f"xlsx={'OK' if gfn_rules_xlsx.exists() else 'MISSING'},",
            f"pickle={'OK' if gfn_rules_pickle.exists() else 'MISSING'},",
            f"rule_order={'OK' if gfn_rule_order.exists() else 'MISSING'},",
            f"checkpoint={'OK' if has_gfn_checkpoint else 'MISSING'}",
        )
        _run_module(project, config_path, "pipelines.stage4_select_rules_gflownet")
        has_gfn_checkpoint = (
            gfn_diverse_checkpoint.exists()
            or gfn_converged_checkpoint.exists()
        )
        if (
            not gfn_rules_xlsx.exists()
            or not gfn_rules_pickle.exists()
            or not gfn_rule_order.exists()
            or not has_gfn_checkpoint
        ):
            raise FileNotFoundError(
                "Stage 4 chạy xong nhưng thiếu rules/rule-order/checkpoint."
            )

    budget = _xlsx_row_count(gfn_rules_xlsx)
    if budget <= 0:
        raise RuntimeError("GFlowNet không chọn luật nào.")
    print(f"Matched budget K={budget}", flush=True)

    ranking_dir = output_dir / "04_filtered_rules"
    ranking_csv = ranking_dir / "rule_ranking_analysis.csv"
    ranking_metrics = (
        ranking_dir / "rule_ranking_analysis_metrics.csv"
    )
    ranking_summary = (
        ranking_dir / "rule_ranking_analysis_summary.txt"
    )
    if (
        ranking_csv.exists()
        and ranking_metrics.exists()
        and ranking_summary.exists()
    ):
        print("SKIP Stage 4b ranking analysis:", ranking_csv)
    else:
        print("RESUME Stage 4b ranking analysis.")
        _run_module(
            project,
            config_path,
            "pipelines.stage4b_analyze_rule_rankings",
        )
        if not all(
            path.exists()
            for path in (
                ranking_csv,
                ranking_metrics,
                ranking_summary,
            )
        ):
            raise FileNotFoundError(
                "Stage 4b không tạo đủ CSV/metrics/summary."
            )

    ranking_output_csv = (
        working_dir / f"seed_{args.seed}_rule_ranking_analysis.csv"
    )
    ranking_output_metrics = (
        working_dir
        / f"seed_{args.seed}_rule_ranking_analysis_metrics.csv"
    )
    ranking_output_summary = (
        working_dir
        / f"seed_{args.seed}_rule_ranking_analysis_summary.txt"
    )
    shutil.copy2(ranking_csv, ranking_output_csv)
    shutil.copy2(ranking_metrics, ranking_output_metrics)
    shutil.copy2(ranking_summary, ranking_output_summary)

    checkpoint_eval_status = (
        gfn_rules_dir / "checkpoint_posterior_evaluation_status.json"
    )
    checkpoint_eval_repeats = (
        gfn_rules_dir / "checkpoint_posterior_evaluation_repeats.csv"
    )
    checkpoint_eval_summary_csv = (
        gfn_rules_dir / "checkpoint_posterior_evaluation_summary.csv"
    )
    checkpoint_eval_summary_json = (
        gfn_rules_dir / "checkpoint_posterior_evaluation_summary.json"
    )
    checkpoint_eval_summary_text = (
        gfn_rules_dir / "checkpoint_posterior_evaluation_summary.txt"
    )
    checkpoint_eval_current = False
    if checkpoint_eval_status.exists():
        status = json.loads(
            checkpoint_eval_status.read_text(encoding="utf-8")
        )
        if status.get("comparison_available"):
            checkpoint_eval_current = (
                status.get("repeats") == args.checkpoint_eval_repeats
                and status.get("samples_per_repeat")
                == args.checkpoint_eval_samples
                and all(
                    path.exists()
                    for path in (
                        checkpoint_eval_repeats,
                        checkpoint_eval_summary_csv,
                        checkpoint_eval_summary_json,
                        checkpoint_eval_summary_text,
                    )
                )
            )
        else:
            # Trạng thái unavailable chỉ còn hợp lệ nếu hiện vẫn thiếu một
            # trong hai checkpoint.
            checkpoint_eval_current = not (
                gfn_diverse_checkpoint.exists()
                and gfn_converged_checkpoint.exists()
            )
    if checkpoint_eval_current:
        print("SKIP Stage 4c checkpoint posterior evaluation.")
    else:
        print(
            "RESUME Stage 4c checkpoint posterior evaluation:",
            f"{args.checkpoint_eval_repeats} x "
            f"{args.checkpoint_eval_samples} samples/checkpoint",
        )
        _run_module(
            project,
            config_path,
            "pipelines.stage4c_evaluate_gflownet_checkpoints",
            "--repeats",
            args.checkpoint_eval_repeats,
            "--samples-per-repeat",
            args.checkpoint_eval_samples,
            "--sample-batch-size",
            args.checkpoint_eval_batch_size,
        )
    if not checkpoint_eval_status.exists():
        raise FileNotFoundError("Stage 4c không tạo status JSON.")

    checkpoint_eval_outputs = []
    status = json.loads(
        checkpoint_eval_status.read_text(encoding="utf-8")
    )
    status_destination = (
        working_dir
        / f"seed_{args.seed}_checkpoint_posterior_evaluation_status.json"
    )
    shutil.copy2(checkpoint_eval_status, status_destination)
    checkpoint_eval_outputs.append(status_destination)
    if status.get("comparison_available"):
        for source, suffix in (
            (checkpoint_eval_repeats, "repeats.csv"),
            (checkpoint_eval_summary_csv, "summary.csv"),
            (checkpoint_eval_summary_json, "summary.json"),
            (checkpoint_eval_summary_text, "summary.txt"),
        ):
            destination = (
                working_dir
                / f"seed_{args.seed}_checkpoint_posterior_evaluation_{suffix}"
            )
            shutil.copy2(source, destination)
            checkpoint_eval_outputs.append(destination)
    _write_config(
        config_path,
        seed=args.seed,
        backbone=args.backbone,
        data_dir=data_dir,
        output_dir=output_dir,
        selection_budget=budget,
    )
    if args.include_bayesian:
        bayesian_config = yaml.safe_load(
            config_path.read_text(encoding="utf-8")
        )
        bayesian_config.setdefault("rule_penalty_bayesian", {})["K"] = (
            args.bayesian_mc_samples
        )
        config_path.write_text(
            yaml.safe_dump(
                bayesian_config,
                sort_keys=False,
                allow_unicode=True,
            ),
            encoding="utf-8",
        )

    gfn_model_dir = output_dir / "05_rules_model"
    gfn_model_checkpoint = (
        gfn_model_dir / "rule_regularized_best.pth"
    )
    gfn_report = _first_report(gfn_model_dir)
    if gfn_report is not None and gfn_model_checkpoint.exists():
        print("SKIP Stage 5 GFlowNet:", gfn_report)
    else:
        _run_module(project, config_path, "pipelines.stage5_train_rule_regularized")
        gfn_report = _first_report(output_dir / "05_rules_model")
        if gfn_report is None:
            raise FileNotFoundError("Stage 5 hoàn tất nhưng không tạo classification report.")

    # Resume chi tiết từng heuristic. Script heuristic ghi artefact của mỗi
    # phương pháp vào thư mục riêng, vì vậy một method hoàn tất không cần chạy
    # lại khi method khác bị gián đoạn.
    heuristic_reports: Dict[str, Path] = {}
    for method in HEURISTICS:
        method_rules = (
            output_dir
            / f"04_filtered_rules_{method}"
            / "selected_rules.xlsx"
        )
        method_report = _first_report(
            output_dir / f"05_rules_model_{method}"
        )
        method_checkpoint = (
            output_dir
            / f"05_rules_model_{method}"
            / "rule_regularized_best.pth"
        )
        if (
            method_rules.exists()
            and method_report is not None
            and method_checkpoint.exists()
        ):
            print(f"SKIP heuristic {method}: {method_report}")
            heuristic_reports[method] = method_report
            continue

        print(
            f"RESUME heuristic {method}: "
            f"rules={'OK' if method_rules.exists() else 'MISSING'}, "
            f"report={'OK' if method_report is not None else 'MISSING'}, "
            f"checkpoint={'OK' if method_checkpoint.exists() else 'MISSING'}"
        )
        _run_module(
            project,
            config_path,
            "pipelines.stage5_train_rule_regularized_heuristics",
            "--methods",
            method,
            "--random_seed",
            args.seed,
        )
        method_report = _first_report(
            output_dir / f"05_rules_model_{method}"
        )
        if (
            not method_rules.exists()
            or method_report is None
            or not method_checkpoint.exists()
        ):
            raise FileNotFoundError(
                f"Heuristic {method} chạy xong nhưng thiếu rules/report."
            )
        heuristic_reports[method] = method_report

    bayesian_report: Optional[Path] = None
    if args.include_bayesian:
        bayesian_dir = output_dir / "05b_rules_model_bayesian"
        bayesian_report = _restore_bayesian_if_available(
            input_root,
            output_dir,
            args.seed,
            args.backbone,
            dataset_id,
        )
        bayesian_checkpoint = (
            bayesian_dir / "rule_regularized_best.pth"
        )
        if (
            bayesian_report is not None
            and bayesian_checkpoint.exists()
        ):
            print(
                "SKIP Bayesian Stage 5: đã tìm thấy report hoàn tất",
                bayesian_report,
            )
        else:
            if bayesian_report is not None:
                print(
                    "Bayesian report cũ tồn tại nhưng thiếu checkpoint; "
                    "cần train lại để tính metric chính xác."
                )
            print(
                "RESUME Bayesian Stage 5:",
                f"MC K={args.bayesian_mc_samples}",
            )
            _run_module(
                project,
                config_path,
                "pipelines.stage5_train_rule_bayesian",
            )
            bayesian_report = _first_report(bayesian_dir)
            if bayesian_report is None:
                raise FileNotFoundError(
                    "Bayesian Stage 5 chạy xong nhưng thiếu report."
                )

    # Metric canonical: đánh giá lại checkpoint trên test loader tái lập.
    # Không dùng các số đã làm tròn trong classification report văn bản.
    exact_args: List[object] = []
    if args.include_bayesian:
        exact_args.append("--include-bayesian")
    _run_module(
        project,
        config_path,
        "pipelines.evaluate_saved_checkpoints",
        *exact_args,
    )
    exact_csv_source = output_dir / "exact_test_metrics_summary.csv"
    exact_json_source = (
        output_dir / "exact_test_metrics_all_methods.json"
    )
    if not exact_csv_source.exists() or not exact_json_source.exists():
        raise FileNotFoundError(
            "Đánh giá checkpoint không tạo đủ exact metrics CSV/JSON."
        )
    exact_csv_path = (
        working_dir / f"seed_{args.seed}_exact_test_metrics.csv"
    )
    exact_json_path = (
        working_dir / f"seed_{args.seed}_exact_test_metrics.json"
    )
    shutil.copy2(exact_csv_source, exact_csv_path)
    shutil.copy2(exact_json_source, exact_json_path)

    counts = {"gflownet": budget}
    for method in HEURISTICS:
        counts[method] = _xlsx_row_count(
            output_dir / f"04_filtered_rules_{method}" / "selected_rules.xlsx"
        )
        if counts[method] != budget:
            raise RuntimeError(
                f"Vi phạm matched-budget seed {args.seed}: "
                f"{method}={counts[method]}, GFlowNet={budget}."
            )

    fairness_path = working_dir / f"seed_{args.seed}_fairness.csv"
    _write_csv(
        fairness_path,
        [
            {
                "seed": args.seed,
                "backbone": args.backbone,
                **counts,
                "matched_budget_pass": True,
            }
        ],
        [
            "seed",
            "backbone",
            "gflownet",
            "random",
            "topk_confidence",
            "greedy_coverage",
            "matched_budget_pass",
        ],
    )

    # Đánh giá trực tiếp chất lượng tập luật bằng đúng RuleSetReward và
    # validation tensors Stage 4 đã dùng. Chạy sau matched-budget audit để
    # bảo đảm bốn phương pháp được so sánh cùng số luật.
    _run_module(
        project,
        config_path,
        "pipelines.analyze_rule_set_quality",
    )
    quality_csv_source = output_dir / "rule_set_quality_comparison.csv"
    quality_json_source = output_dir / "rule_set_quality_comparison.json"
    if not quality_csv_source.exists() or not quality_json_source.exists():
        raise FileNotFoundError(
            "Phân tích chất lượng luật không tạo đủ CSV/JSON."
        )
    quality_csv_path = (
        working_dir / f"seed_{args.seed}_rule_set_quality.csv"
    )
    quality_json_path = (
        working_dir / f"seed_{args.seed}_rule_set_quality.json"
    )
    shutil.copy2(quality_csv_source, quality_csv_path)
    shutil.copy2(quality_json_source, quality_json_path)

    with exact_csv_source.open(newline="", encoding="utf-8") as file:
        exact_rows = list(csv.DictReader(file))
    exact_by_method = {row["method"]: row for row in exact_rows}
    expected_exact_methods = {
        "cnn_baseline",
        "gflownet_db",
        *HEURISTICS,
    }
    if args.include_bayesian:
        expected_exact_methods.add("gflownet_db_bayesian")
    missing_exact = expected_exact_methods - set(exact_by_method)
    if missing_exact:
        raise ValueError(
            "Exact metrics thiếu method: "
            + ", ".join(sorted(missing_exact))
        )

    rule_counts = {
        "cnn_baseline": 0,
        "gflownet_db": budget,
        **counts,
        "gflownet_db_bayesian": "",
    }
    result_rows = []
    for method in [
        "cnn_baseline",
        "gflownet_db",
        *HEURISTICS,
        *(["gflownet_db_bayesian"] if args.include_bayesian else []),
    ]:
        exact = exact_by_method[method]
        result_rows.append(
            {
                "seed": args.seed,
                "backbone": args.backbone,
                "method": method,
                "n_rules_selected": rule_counts[method],
                "test_accuracy": float(exact["test_accuracy"]),
                "test_macro_precision": float(
                    exact["test_macro_precision"]
                ),
                "test_macro_recall": float(
                    exact["test_macro_recall"]
                ),
                "test_f1_macro": float(exact["test_f1_macro"]),
                "test_weighted_f1": float(exact["test_weighted_f1"]),
            }
        )

    heuristic_comparison_rows = []
    for method in HEURISTICS:
        exact = exact_by_method[method]
        heuristic_comparison_rows.append(
            {
                "method": method,
                "n_rules_selected": counts[method],
                "test_accuracy": float(exact["test_accuracy"]),
                "test_f1_macro": float(exact["test_f1_macro"]),
                "save_dir": str(output_dir / f"05_rules_model_{method}"),
            }
        )

    # Mỗi lần gọi script cho một method sẽ ghi đè CSV tạm. Dựng lại bảng
    # canonical từ ba report riêng sau khi tất cả method đã hoàn tất.
    heuristic_csv = output_dir / "rule_selection_finetune_comparison.csv"
    _write_csv(
        heuristic_csv,
        heuristic_comparison_rows,
        [
            "method",
            "n_rules_selected",
            "test_accuracy",
            "test_f1_macro",
            "save_dir",
        ],
    )

    results_path = working_dir / f"seed_{args.seed}_results.csv"
    _write_csv(
        results_path,
        result_rows,
        [
            "seed",
            "backbone",
            "method",
            "n_rules_selected",
            "test_accuracy",
            "test_macro_precision",
            "test_macro_recall",
            "test_f1_macro",
            "test_weighted_f1",
        ],
    )
    runtime_summary_path, runtime_events_path = _write_runtime_summary(
        output_dir,
        working_dir,
        seed=args.seed,
        backbone=args.backbone,
        include_bayesian=args.include_bayesian,
    )
    export_dir = working_dir / "experiment_exports" / run_id
    export_dir.mkdir(parents=True, exist_ok=True)
    lightweight_exports = {
        "results.csv": results_path,
        "fairness.csv": fairness_path,
        "exact_test_metrics.csv": exact_csv_path,
        "exact_test_metrics.json": exact_json_path,
        "rule_set_quality.csv": quality_csv_path,
        "rule_set_quality.json": quality_json_path,
        "rule_ranking_analysis.csv": ranking_output_csv,
        "rule_ranking_metrics.csv": ranking_output_metrics,
        "rule_ranking_summary.txt": ranking_output_summary,
        "runtime_summary.csv": runtime_summary_path,
        "runtime_events.json": runtime_events_path,
    }
    for export_name, source in lightweight_exports.items():
        shutil.copy2(source, export_dir / export_name)
    for checkpoint_output in checkpoint_eval_outputs:
        shutil.copy2(checkpoint_output, export_dir / checkpoint_output.name)

    run_manifest.update(
        {
            "status": "complete",
            "finished_at_utc": datetime.now(timezone.utc).isoformat(),
            "output_dir": str(output_dir),
            "managed_export_dir": str(export_dir),
            "result_rows": len(result_rows),
        }
    )
    _write_run_manifest(manifest_path, run_manifest)
    _write_run_manifest(export_dir / "run_manifest.json", run_manifest)
    archive = shutil.make_archive(
        str(working_dir / f"seed_{args.seed}_artifacts"),
        "gztar",
        root_dir=output_dir,
    )
    print("\nHoàn tất:")
    print(" -", results_path)
    print(" -", fairness_path)
    print(" -", exact_csv_path)
    print(" -", exact_json_path)
    print(" -", quality_csv_path)
    print(" -", quality_json_path)
    print(" -", ranking_output_csv)
    print(" -", ranking_output_metrics)
    print(" -", ranking_output_summary)
    print(" -", runtime_summary_path)
    print(" -", runtime_events_path)
    print(" -", export_dir / "run_manifest.json")
    print(" -", export_dir)
    for checkpoint_output in checkpoint_eval_outputs:
        print(" -", checkpoint_output)
    print(" -", archive)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="params.yaml")
    parser.add_argument("--seed", type=int, required=True)
    parser.add_argument("--dataset-id", default="default_dataset")
    parser.add_argument("--account-id", default="unassigned")
    parser.add_argument("--run-id", default=None)
    parser.add_argument(
        "--backbone",
        default="mobilenetv3_small",
        choices=SUPPORTED_BACKBONES,
        help="Backbone CNN/Transformer dùng cho toàn bộ pipeline của seed.",
    )
    parser.add_argument("--data-dir", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--project-dir", default=os.getcwd())
    parser.add_argument("--working-dir", default="/kaggle/working")
    parser.add_argument(
        "--kaggle-input-root",
        default="/kaggle/input",
        help="Thư mục Add Input. Module tự restore output cùng seed nếu tìm thấy.",
    )
    parser.add_argument(
        "--include-bayesian",
        action="store_true",
        help="Chạy/resume Bayesian Stage 5 sau ba heuristic.",
    )
    parser.add_argument(
        "--bayesian-mc-samples",
        type=int,
        default=32,
        help="Số ruleset Monte Carlo mỗi bước Bayesian Stage 5.",
    )
    parser.add_argument(
        "--checkpoint-eval-repeats",
        type=int,
        default=5,
    )
    parser.add_argument(
        "--checkpoint-eval-samples",
        type=int,
        default=1000,
    )
    parser.add_argument(
        "--checkpoint-eval-batch-size",
        type=int,
        default=250,
    )
    return parser


if __name__ == "__main__":
    parsed_args = build_parser().parse_args()
    try:
        run(parsed_args)
    except Exception as error:
        failed_manifest_path = Path(parsed_args.output_dir) / "run_manifest.json"
        if failed_manifest_path.exists():
            failed_manifest = json.loads(
                failed_manifest_path.read_text(encoding="utf-8")
            )
            failed_manifest.update(
                {
                    "status": "failed",
                    "finished_at_utc": datetime.now(
                        timezone.utc
                    ).isoformat(),
                    "error_type": type(error).__name__,
                    "error_message": str(error),
                }
            )
            _write_run_manifest(failed_manifest_path, failed_manifest)
        raise
